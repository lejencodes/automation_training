# import xlrd

from util.log import LogMessage
import openpyxl

logs = LogMessage()


def get_test_data(filepath, shtname):  # sourcery skip: merge-dict-assign
    try:
        file = openpyxl.load_workbook(filepath)
        sht = file[shtname]
        nrows = sht.max_row
        listdata = []
        for i in range(2, nrows+1):
            dict_canshu = {'id': sht.cell(i, 1).value}
            dict_canshu.update(eval(sht.cell(i, 3).value))
            dict_canshu.update(eval(sht.cell(i, 4).value))
            listdata.append(dict_canshu)
        return listdata
    except Exception as e:
        logs.error_log('Unable to get test data，Reason：%s' % e)


# import os

# print(get_test_data(os.path.join('data', 'case.xlsx'), 'Login'))
